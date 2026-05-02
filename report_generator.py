"""
PDF and Excel report generation for the Business Analytics Platform.
"""

import io
import logging
from datetime import datetime
from typing import Any, Dict, List, Optional

import pandas as pd

from config import APP_TITLE, APP_VERSION
from utils import format_currency, format_percentage, logger

# ─── Optional heavy imports ───────────────────────────────────────────────────

try:
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        HRFlowable,
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
    )

    _REPORTLAB_AVAILABLE = True
except ImportError:
    _REPORTLAB_AVAILABLE = False
    logger.warning("reportlab not installed – PDF export disabled.")

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils.dataframe import dataframe_to_rows

    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed – Excel export disabled.")


# ─── PDF Report ───────────────────────────────────────────────────────────────


def generate_pdf_report(
    kpis: Dict[str, Any],
    df: pd.DataFrame,
    ai_insights: Optional[Dict[str, str]] = None,
    company_name: str = "Business Analytics Report",
) -> Optional[bytes]:
    """
    Generate a styled PDF report and return it as bytes.
    Returns None if reportlab is not available.
    """
    if not _REPORTLAB_AVAILABLE:
        return None

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()

    # Custom styles
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Title"],
        fontSize=24,
        textColor=rl_colors.HexColor("#6C63FF"),
        spaceAfter=6,
    )
    subtitle_style = ParagraphStyle(
        "CustomSubtitle",
        parent=styles["Normal"],
        fontSize=11,
        textColor=rl_colors.HexColor("#A0A0C0"),
        spaceAfter=20,
    )
    section_style = ParagraphStyle(
        "SectionHeader",
        parent=styles["Heading2"],
        fontSize=14,
        textColor=rl_colors.HexColor("#6C63FF"),
        spaceBefore=16,
        spaceAfter=8,
    )
    body_style = ParagraphStyle(
        "Body",
        parent=styles["Normal"],
        fontSize=10,
        textColor=rl_colors.HexColor("#333333"),
        spaceAfter=8,
        leading=16,
    )

    story = []

    # Header
    story.append(Paragraph(APP_TITLE, title_style))
    story.append(Paragraph(company_name, subtitle_style))
    story.append(Paragraph(
        f"Generated: {datetime.now().strftime('%B %d, %Y at %H:%M UTC')} | Version {APP_VERSION}",
        body_style,
    ))
    story.append(HRFlowable(width="100%", thickness=2, color=rl_colors.HexColor("#6C63FF")))
    story.append(Spacer(1, 0.5 * cm))

    # KPI Section
    story.append(Paragraph("Key Performance Indicators", section_style))

    kpi_data = [["Metric", "Value", "Status"]]
    kpi_map = {
        "Total Revenue": ("total_revenue", "$"),
        "Total Expenses": ("total_expenses", "$"),
        "Net Profit": ("net_profit", "$"),
        "Profit Margin": ("profit_margin", "%"),
        "Revenue Growth": ("revenue_growth", "%"),
        "Expense Growth": ("expense_growth", "%"),
    }

    for label, (key, sym) in kpi_map.items():
        val = kpis.get(key)
        if val is not None:
            if sym == "$":
                formatted = format_currency(val)
            else:
                formatted = format_percentage(val)
            status = "✓" if (val >= 0 if key != "total_expenses" else True) else "✗"
            kpi_data.append([label, formatted, status])

    if len(kpi_data) > 1:
        kpi_table = Table(kpi_data, colWidths=[6 * cm, 5 * cm, 3 * cm])
        kpi_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#6C63FF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#F8F8FF")]),
            ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#DDDDDD")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 0.5 * cm))

    # Data Summary
    story.append(Paragraph("Data Summary", section_style))
    story.append(Paragraph(
        f"Total Records: {len(df):,} | Columns: {len(df.columns)} | "
        f"Memory Usage: {df.memory_usage(deep=True).sum() / 1024:.1f} KB",
        body_style,
    ))

    # Numeric stats table
    numeric_df = df.select_dtypes(include="number")
    if not numeric_df.empty:
        stats = numeric_df.describe().round(2)
        stats_data = [["Metric"] + stats.columns.tolist()]
        for idx in stats.index:
            row = [str(idx)] + [str(stats.loc[idx, col]) for col in stats.columns]
            stats_data.append(row)

        stats_table = Table(stats_data)
        stats_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#3ECFCF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, rl_colors.HexColor("#DDDDDD")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.white, rl_colors.HexColor("#F0FFFE")]),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(stats_table)
        story.append(Spacer(1, 0.5 * cm))

    # AI Insights
    if ai_insights:
        for section_key, section_title in [
            ("executive_summary", "Executive Summary"),
            ("trend_analysis", "Trend Analysis"),
            ("recommendations", "Strategic Recommendations"),
            ("anomaly_explanation", "Anomaly Analysis"),
            ("comparative_analysis", "Comparative Analysis"),
        ]:
            content = ai_insights.get(section_key)
            if content and not content.startswith("⚠️"):
                story.append(Paragraph(section_title, section_style))
                # Clean markdown-style formatting for PDF
                clean_content = (
                    content
                    .replace("**", "")
                    .replace("##", "")
                    .replace("# ", "")
                    .replace("*", "•")
                )
                for line in clean_content.split("\n"):
                    line = line.strip()
                    if line:
                        story.append(Paragraph(line, body_style))
                story.append(Spacer(1, 0.3 * cm))

    # Footer
    story.append(HRFlowable(width="100%", thickness=1, color=rl_colors.HexColor("#DDDDDD")))
    story.append(Spacer(1, 0.2 * cm))
    story.append(Paragraph(
        f"© {datetime.now().year} {APP_TITLE} | Confidential Business Report",
        ParagraphStyle("Footer", parent=styles["Normal"], fontSize=8,
                       textColor=rl_colors.HexColor("#999999"), alignment=1),
    ))

    doc.build(story)
    return buffer.getvalue()


# ─── Excel Export ─────────────────────────────────────────────────────────────


def generate_excel_report(
    df: pd.DataFrame,
    kpis: Dict[str, Any],
    pl_df: Optional[pd.DataFrame] = None,
    yoy_df: Optional[pd.DataFrame] = None,
    company_name: str = "Business Report",
) -> Optional[bytes]:
    """
    Generate a formatted Excel workbook and return it as bytes.
    Returns None if openpyxl is not available.
    """
    if not _OPENPYXL_AVAILABLE:
        return None

    buffer = io.BytesIO()
    wb = openpyxl.Workbook()

    # ── KPI Sheet ─────────────────────────────────────────────────────────────
    ws_kpi = wb.active
    ws_kpi.title = "KPIs"

    _xl_header(ws_kpi, f"{company_name} – Key Performance Indicators", cols=3)

    headers = ["Metric", "Value", "Growth vs Previous"]
    _xl_row(ws_kpi, headers, bold=True, bg="#6C63FF", fg="FFFFFF")

    kpi_rows = [
        ("Total Revenue", format_currency(kpis.get("total_revenue") or 0), format_percentage(kpis.get("revenue_growth") or 0)),
        ("Total Expenses", format_currency(kpis.get("total_expenses") or 0), format_percentage(kpis.get("expense_growth") or 0)),
        ("Net Profit", format_currency(kpis.get("net_profit") or 0), format_percentage(kpis.get("profit_growth") or 0)),
        ("Profit Margin", format_percentage(kpis.get("profit_margin") or 0), "—"),
        ("Record Count", f"{kpis.get('record_count', 0):,}", "—"),
    ]
    for i, row in enumerate(kpi_rows):
        bg = "F8F8FF" if i % 2 == 0 else "FFFFFF"
        _xl_row(ws_kpi, list(row), bg=bg)

    _xl_autofit(ws_kpi)

    # ── Raw Data Sheet ────────────────────────────────────────────────────────
    ws_data = wb.create_sheet("Raw Data")
    _xl_header(ws_data, "Raw Data", cols=len(df.columns))
    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)
    # Style header row (row 3 because of title + blank)
    for cell in ws_data[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="3ECFCF")
    _xl_autofit(ws_data)

    # ── P&L Sheet ─────────────────────────────────────────────────────────────
    if pl_df is not None and not pl_df.empty:
        ws_pl = wb.create_sheet("P&L Analysis")
        _xl_header(ws_pl, "Profit & Loss Analysis", cols=len(pl_df.columns))
        for r in dataframe_to_rows(pl_df, index=False, header=True):
            ws_pl.append(r)
        for cell in ws_pl[3]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="FF6584")
        _xl_autofit(ws_pl)

    # ── YoY Sheet ─────────────────────────────────────────────────────────────
    if yoy_df is not None and not yoy_df.empty:
        ws_yoy = wb.create_sheet("YoY Comparison")
        _xl_header(ws_yoy, "Year-over-Year Comparison", cols=len(yoy_df.columns))
        for r in dataframe_to_rows(yoy_df, index=False, header=True):
            ws_yoy.append(r)
        for cell in ws_yoy[3]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="FFBF69")
        _xl_autofit(ws_yoy)

    # ── Stats Sheet ───────────────────────────────────────────────────────────
    ws_stats = wb.create_sheet("Statistics")
    _xl_header(ws_stats, "Descriptive Statistics", cols=len(df.select_dtypes(include="number").columns) + 1)
    stats_df = df.select_dtypes(include="number").describe().round(4).reset_index()
    stats_df.rename(columns={"index": "Statistic"}, inplace=True)
    for r in dataframe_to_rows(stats_df, index=False, header=True):
        ws_stats.append(r)
    for cell in ws_stats[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="A78BFA")
    _xl_autofit(ws_stats)

    wb.save(buffer)
    return buffer.getvalue()


# ─── Excel Helpers ────────────────────────────────────────────────────────────


def _xl_header(ws, title: str, cols: int) -> None:
    """Add a merged title header row to an Excel worksheet."""
    ws.append([title])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(cols, 1))
    cell = ws.cell(row=1, column=1)
    cell.font = Font(bold=True, size=14, color="6C63FF")
    cell.alignment = Alignment(horizontal="center")
    ws.append([])  # blank row


def _xl_row(ws, values: list, bold: bool = False, bg: str = "", fg: str = "000000") -> None:
    """Append a styled row to an Excel worksheet."""
    ws.append(values)
    row_idx = ws.max_row
    # Strip '#' from hex colors for openpyxl
    fg_hex = fg.lstrip("#")
    bg_hex = bg.lstrip("#") if bg else ""
    for cell in ws[row_idx]:
        cell.font = Font(bold=bold, color=fg_hex)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if bg_hex:
            cell.fill = PatternFill("solid", fgColor=bg_hex)


def _xl_autofit(ws) -> None:
    """Auto-fit column widths in an Excel worksheet."""
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            try:
                if hasattr(cell, "column_letter"):
                    col_letter = cell.column_letter
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


# ─── CSV Export ───────────────────────────────────────────────────────────────


def dataframe_to_csv(df: pd.DataFrame) -> bytes:
    """Convert a DataFrame to CSV bytes."""
    return df.to_csv(index=False).encode("utf-8")
